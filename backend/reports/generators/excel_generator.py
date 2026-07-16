"""
Excel report generator using openpyxl.
Creates a formatted .xlsx file from a transaction queryset.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel_report(transactions):
    """
    Return a BytesIO buffer containing an .xlsx workbook.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # ── Styles ─────────────────────────────────────────────────────────────
    header_fill = PatternFill(fill_type='solid', fgColor='6366F1')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')

    income_fill  = PatternFill(fill_type='solid', fgColor='DCFCE7')
    expense_fill = PatternFill(fill_type='solid', fgColor='FEE2E2')

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # ── Header Row ─────────────────────────────────────────────────────────
    headers = ['Date', 'Type', 'Category', 'Amount (₹)', 'Description']
    col_widths = [14, 12, 20, 16, 40]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell            = ws.cell(row=1, column=col_idx, value=header)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center_align
        cell.border     = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25

    # ── Data Rows ──────────────────────────────────────────────────────────
    for row_idx, txn in enumerate(transactions, 2):
        row_fill = income_fill if txn.transaction_type == 'income' else expense_fill
        data = [
            str(txn.date),
            txn.transaction_type.capitalize(),
            txn.category.name if txn.category else 'N/A',
            float(txn.amount),
            txn.description,
        ]
        for col_idx, value in enumerate(data, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = Alignment(vertical='center')

    # ── Auto-filter ────────────────────────────────────────────────────────
    ws.auto_filter.ref = f'A1:E{ws.max_row}'
    ws.freeze_panes    = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
